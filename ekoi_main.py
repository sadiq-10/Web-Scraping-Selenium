
import openpyxl as xl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# create a csv sheet for saving data
# with open('ekoi_main_site.csv', 'w') as file:
#     file.write("title_text; category; images; size_list; reguler_price; "
#                "discounted_price; short_description; description /n")

workbook = xl.Workbook()
sheet = workbook.active
workbook.save('ekoi_main_site.xlsx')
xlSheetRow = 1

PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)
driver.maximize_window()



# url = driver.get('https://www.ekoi.com/en-au/')
# category_list = driver.find_element_by_id('nav-main-ekoimenu')
# sub_cat_list = category_list.find_elements_by_class_name('ekoimenu_li_level0')
# for sub_cat in sub_cat_list:
#     sub_sub_cat = sub_cat.find_elements_by_class_name('ekoimenu_category_block_heading')
#     for s in sub_sub_cat:
#         print(s.get_attribute('href'))


# get_elements(get_str('https://www.ekoi.com/en-bd/1807-summer-clothing'))

category_list = ['https://www.ekoi.com/en-bd/32-helmet-caps',
'https://www.ekoi.com/en-bd/1835-bandeaux-hiver',
'https://www.ekoi.com/en-bd/1543-accessoires-chauffants',
'https://www.ekoi.com/en-bd/37-water-bottles',
'https://www.ekoi.com/en-bd/294-bottle-cages',
'https://www.ekoi.com/en-bd/842-grips-and-sports-bags',
'https://www.ekoi.com/en-bd/841-travel-bags',
'https://www.ekoi.com/en-bd/843-bike-bags',
'https://www.ekoi.com/en-bd/185-wheel-bags',
'https://www.ekoi.com/en-bd/421-long-sleeves-jersey',
'https://www.ekoi.com/en-bd/420-short-sleeve-mtb-jerseys',
'https://www.ekoi.com/en-bd/1669-pantalons-vtt',
'https://www.ekoi.com/en-bd/425-mtb-gloves',
'https://www.ekoi.com/en/532-vetements-de-protection-velo?f_discipline_3=113',
'https://www.ekoi.com/en-bd/1267-all-mountain-helmets',
'https://www.ekoi.com/en-bd/1723-all-mountain-shoes',
'https://www.ekoi.com/en-bd/1243-xc-shoes',
'https://www.ekoi.com/en-bd/1883-protections-vtt',
'https://www.ekoi.com/en-bd/1722-xc-helmets']

for link in category_list:
    # for getting all products link in a sub-category
    driver.get(link)
    driver.implicitly_wait(20)

    all_product = []
    product_list = driver.find_element_by_class_name('es-productslist')
    list_of_links = product_list.find_elements_by_class_name('es-product__title')
    for l in list_of_links:
        link_id = l.get_attribute('href')
        all_product.append(link_id)

    for links in all_product:
        column_arr = []
        xl_sheet_colmn = 1
        driver.get(links)
        driver.implicitly_wait(5)
         # print title
        try:
            title_element = driver.find_element_by_xpath('//div[@class="col-xs-12 col-sm-8"]')
            title_text = title_element.find_element_by_tag_name('h1').text
            print(title_text)
        except:
            print("no need")
            pass

        column_arr.append(title_text)


        # print category
        category = ''
        try:
            cat_div = driver.find_element_by_xpath('//div[@class="breadcrumb-container"]')
            cat_list = cat_div.find_elements_by_xpath('//*[@id="content"]/div[1]/ul/span/a')
            for sp in cat_list:
                type_of_product = sp.text
                category += type_of_product + '/'
            print(category)
        except:
            print("no need")
            pass
        column_arr.append(category)
    #
    #
    #     # print images of product
        images = ''
        try:
            img_element_div = driver.find_element_by_xpath('//div[@class="thumbnails col-xs-12 hidden-md hidden-lg"]')
            img_element_li = img_element_div.find_element_by_tag_name('ul').find_elements_by_tag_name('li')
            for img in img_element_li:
                img_list = img.find_element_by_tag_name('a').get_attribute('href')
                images += img_list + ", "
            # print(images)
        except:
            print("no need")
        column_arr.append(images)
    #
    #
    #
    #
    #     # size chart
        try:
            size_list = ''
            size_div = driver.find_element_by_xpath('//ul[@class="sizeSelectButtons sizeSelectButtons_8"]').find_elements_by_tag_name('li')
            for s in size_div:
                size = s.get_attribute("innerHTML")
                size_list += size + ", "
            column_arr.append(size_list)
        except:
            print("No need")
            pass

    #
    #
    #     # price & discount price
        try:
            discounted_price = driver.find_element_by_id('our_price_display').text
        except:
            print("empty")

        column_arr.append(discounted_price)


        try:
            reguler_price = driver.find_element_by_id('old_price_display').text
        except:
            reguler_price = driver.find_elements_by_class_name('es-product__prices')

        column_arr.append(reguler_price)
    #
    #
    #     #discription & specification
        try:
            discription = driver.find_element_by_id('description_tab').find_element_by_class_name('description-short')
            discription_short = discription.find_elements_by_tag_name('p')
            full_description = ""
            for d in discription_short:
                short_description = d.text
                column_arr.append(short_description)


            feature = driver.find_element_by_id('caracteristiques_tab').find_elements_by_tag_name('p')
            for f in feature:
                description = f.text
                column_arr.append(description)


        except:
            pass

        for c in column_arr:
            if type(c) == str:
                sheet.cell(xlSheetRow, xl_sheet_colmn, c)
            else:
                sheet.cell(xlSheetRow, xl_sheet_colmn)
            xl_sheet_colmn += 1
            print("ROW: " + xlSheetRow.__str__())
            print("COLUMN: " + xl_sheet_colmn.__str__())
        xlSheetRow += 1
        workbook.save('ekoi_main_site.xlsx')



driver.close()
